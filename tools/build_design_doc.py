"""
Solution Design Document generator for:
  "Tariff & Geopolitical Shock Early-Warning System"

Produces a multi-sheet, professionally styled Excel workbook:
  1. Cover
  2. Document Control (revision history + glossary)
  3. Solution Overview (context, scope, goals, assumptions, constraints)
  4. Use Case Diagram (visual) + use case catalog
  5. Data Sources & Access
  6. Model Selection & Rationale
  7. Text-Fusion Architecture (visual data flow) + component catalog
  8. Repository / File Structure
  9. Evaluation & Backtest Design
 10. Non-Functional Requirements & Risks

Run:  python build_design_doc.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ----------------------------------------------------------------------------
# Design tokens (one palette across the whole document)
# ----------------------------------------------------------------------------
NAVY      = "1F3864"   # primary / headers
BLUE      = "2E5496"   # secondary
STEEL     = "3E6DA6"   # accent
LIGHT     = "D9E1F2"   # light row / box fill
LIGHTER   = "EAF0FA"   # zebra
MINT      = "D6E9D5"   # "chosen / good" fill
MINT_DK   = "3F7D3A"
AMBER     = "FCE4B6"   # "consider / caution" fill
AMBER_DK  = "9C6500"
GREY      = "808080"
GREY_LT   = "F2F2F2"
WHITE     = "FFFFFF"
CARD_ACT  = "C9D8EF"   # actor box
CARD_UC   = "FFFFFF"   # use case box
BOUND     = "F4F7FC"   # system boundary fill

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color=NAVY)
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_BOX  = Border(left=med, right=med, top=med, bottom=med)


def F(size=10, bold=False, color="000000", italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def fill(hexcol):
    return PatternFill("solid", fgColor=hexcol)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, subtitle, last_col="H"):
    """Standard band at the top of a content sheet."""
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = F(16, True, WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last_col}2")
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = F(10, False, WHITE, italic=True)
    c2.fill = fill(BLUE)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def table_header(ws, row, headers, start_col=1, fillcol=NAVY):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F(10, True, WHITE)
        c.fill = fill(fillcol)
        c.alignment = CENTER
        c.border = BORDER_THIN


def table_rows(ws, start_row, rows, start_col=1, zebra=True, valign_top=True,
               fills=None):
    align = LEFT_T if valign_top else LEFT
    for r, rowvals in enumerate(rows):
        excel_row = start_row + r
        rowfill = None
        if fills and fills.get(r) is not None:
            rowfill = fills[r]
        elif zebra and r % 2 == 1:
            rowfill = LIGHTER
        for i, val in enumerate(rowvals):
            c = ws.cell(row=excel_row, column=start_col + i, value=val)
            c.font = F(9.5)
            c.alignment = align
            c.border = BORDER_THIN
            if rowfill:
                c.fill = fill(rowfill)


def box(ws, r1, c1, r2, c2, text, fillcol, fontcol="000000", bold=True,
        size=10, border=BORDER_BOX, italic=False):
    """Draw a merged, filled, bordered box on a fine grid."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.fill = fill(fillcol)
    cell.font = F(size, bold, fontcol, italic=italic)
    cell.alignment = CENTER
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).border = border


def note(ws, r, c, text, size=8.5, color=GREY):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = F(size, False, color, italic=True)
    cell.alignment = LEFT


# ============================================================================
wb = Workbook()

PROJECT = "Tariff & Geopolitical Shock Early-Warning System"
DOCTYPE = "Solution Design Document"

# ----------------------------------------------------------------------------
# SHEET 1 - COVER
# ----------------------------------------------------------------------------
ws = wb.active
ws.title = "1. Cover"
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 3, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 3})

for r in range(1, 40):
    ws.row_dimensions[r].height = 18

box(ws, 3, 2, 6, 6, DOCTYPE, NAVY, WHITE, bold=True, size=22)
box(ws, 7, 2, 9, 6, PROJECT, STEEL, WHITE, bold=True, size=15)
box(ws, 10, 2, 11, 6,
    "Zero-shot foundation-model forecasting  |  Event-informed reasoning  |  Causal break attribution",
    BLUE, WHITE, bold=False, size=10, italic=True)

meta = [
    ("Document Type", DOCTYPE),
    ("Version", "1.4 (v2 delivered)"),
    ("Status", "v1 + v2 complete: two live verticals, alert artifacts, covariates closed-negative"),
    ("Date", "2026-07-11"),
    ("Target Vertical (v1)", "Steel / steel-exposed sectors (HRC futures, SLX proxy)"),
    ("Primary Objective", "Give risk teams early, calibrated, attributable warning of "
                          "tariff- and geopolitics-driven commodity shocks"),
    ("Intended Audience", "Risk analytics, supply-chain risk, engineering reviewers, hiring panels"),
]
rr = 14
box(ws, rr, 2, rr, 6, "Document Summary", BLUE, WHITE, size=11)
rr += 1
for k, v in meta:
    box(ws, rr, 2, rr, 2, k, LIGHT, NAVY, bold=True, size=9.5)
    box(ws, rr, 3, rr, 6, v, WHITE, "000000", bold=False, size=9.5, border=BORDER_THIN)
    rr += 1

box(ws, rr + 1, 2, rr + 1, 6,
    "Design baseline - built on public data only; fully reproducible by a reviewer.",
    GREY_LT, GREY, bold=False, size=8.5, italic=True, border=BORDER_THIN)

# ----------------------------------------------------------------------------
# SHEET 2 - DOCUMENT CONTROL
# ----------------------------------------------------------------------------
ws = wb.create_sheet("2. Document Control")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 12, "B": 16, "C": 46, "D": 16, "E": 14})
title_block(ws, "Document Control", "Revision history, approvals, and glossary", "E")

r = 4
ws.cell(row=r, column=1, value="Revision History").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Version", "Date", "Change Description", "Status"])
rev = [
    ["0.1", "2026-07-06", "Initial skeleton, scope framing", "Draft"],
    ["0.5", "2026-07-07", "Data sources, model trade study, architecture", "Draft"],
    ["1.0", "2026-07-08", "Baseline design; v1 scope frozen to steel vertical", "For Review"],
    ["1.1", "2026-07-09", "v1 backbone set to TimesFM 2.5 (quantile head) + CQR calibration; validated on SLX", "For Review"],
    ["1.2", "2026-07-10", "v1 delivered: ACI+GARCH pipeline live with F1 alert filter; causal attribution closed-rejected (two pre-registered nulls); Build Log rows A-J complete", "v1 Complete"],
    ["1.3", "2026-07-10", "v2 plan logged (sheet 12): W1 covariates via TimesFM XReg (Moirai-2 fallback), W2 aluminum vertical config-only, W3 alert delivery; W1 eval pre-registered", "v2 In Progress"],
    ["1.4", "2026-07-11", "v2 delivered: W1 closed-negative (XReg +22.7% width, does not ship; v1 stands), W2 aluminum vertical live config-only, W3 alert artifacts + inert webhook shipped", "v2 Complete"],
]
table_rows(ws, r + 1, rev)

r = r + 1 + len(rev) + 2
ws.cell(row=r, column=1, value="Approvals").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Role", "Name", "Responsibility", "Sign-off", "Date"])
appr = [
    ["Design Authority", "(owner)", "Overall design sign-off", "Pending", ""],
    ["Data / ML Lead", "(TBD)", "Model + data feasibility", "Pending", ""],
    ["Risk Stakeholder", "(TBD)", "Business value & metrics", "Pending", ""],
]
table_rows(ws, r + 1, appr)

r = r + 1 + len(appr) + 2
ws.cell(row=r, column=1, value="Glossary").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Term", "Definition", "", "", ""])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
gloss = [
    ["Zero-shot", "Forecasting a new series with a pretrained model, no per-series retraining."],
    ["Foundation model", "Large pretrained time-series model (Chronos-2, TimesFM, Moirai-2)."],
    ["Conformal prediction", "Wrapper giving prediction intervals with empirically guaranteed coverage."],
    ["CQR", "Conformalized Quantile Regression; offsets a model's q10/q90 band to hit target coverage."],
    ["TimesFM 2.5", "Google's 200M-param decoder-only time-series foundation model; zero-shot, quantile head."],
    ["Coverage", "Fraction of actuals that fall inside the predicted interval (should match nominal, e.g. 90%)."],
    ["Interval breach", "Actual value falls outside the calibrated forecast interval -> candidate shock."],
    ["Changepoint", "Statistically detected structural break in a series (via PELT / BOCPD)."],
    ["Attribution", "Ranking candidate events to name the most likely driver of a break."],
    ["EWS", "Early-Warning System."],
    ["MASE", "Mean Absolute Scaled Error; scale-free forecast accuracy vs seasonal-naive."],
    ["FBX", "Freightos Baltic Index (container shipping cost)."],
]
for i, (t, d) in enumerate(gloss):
    er = r + 1 + i
    c1 = ws.cell(row=er, column=1, value=t); c1.font = F(9.5, True, NAVY)
    c1.border = BORDER_THIN; c1.alignment = LEFT_T
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=5)
    c2 = ws.cell(row=er, column=2, value=d); c2.font = F(9.5); c2.alignment = LEFT_T
    for cc in range(1, 6):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 6):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 3 - SOLUTION OVERVIEW
# ----------------------------------------------------------------------------
ws = wb.create_sheet("3. Solution Overview")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 3, "B": 26, "C": 66, "D": 3})
title_block(ws, "Solution Overview", "Problem context, scope, goals, assumptions, constraints", "D")

def section(ws, r, heading):
    box(ws, r, 2, r, 3, heading, BLUE, WHITE, size=11)
    return r + 1

def kv_rows(ws, r, pairs):
    for k, v in pairs:
        box(ws, r, 2, r, 2, k, LIGHT, NAVY, bold=True, size=9.5, border=BORDER_THIN)
        cell = ws.cell(row=r, column=3, value=v)
        cell.font = F(9.5); cell.alignment = LEFT_T; cell.border = BORDER_THIN
        ws.row_dimensions[r].height = max(28, 14 * (1 + len(v) // 78))
        r += 1
    return r

r = 4
r = section(ws, r, "Problem Context")
r = kv_rows(ws, r, [
    ("The problem", "Tariff and geopolitical policy shocks move commodity prices, shipping "
                    "costs, and input availability faster than classical models react. Risk and "
                    "procurement teams learn about the driver only after the price has moved."),
    ("Who feels it", "Commodity trading / risk desks, manufacturer procurement, trade-credit "
                     "underwriters, freight & logistics planners."),
    ("Why now", "Foundation time-series models now give strong zero-shot forecasts; the open gap "
                "is calibrated uncertainty plus naming the causal event, not raw accuracy."),
])
r += 1
r = section(ws, r, "Solution Goals")
r = kv_rows(ws, r, [
    ("G1 - Calibrated forecast", "Zero-shot forecast with prediction intervals whose coverage "
                                 "matches nominal (e.g. 90% interval covers ~90%)."),
    ("G2 - Event-aware", "Ingest tariff / policy events and raise a risk flag when a forecast is "
                         "about to be invalidated by an event."),
    ("G3 - Attributable", "When actuals breach the interval, name the most likely driving event "
                          "with a verifiable source link."),
    ("G4 - Checkable", "Every claim is dated and falsifiable; backtested against real 2025-2026 events."),
])
r += 1
r = section(ws, r, "Scope")
r = kv_rows(ws, r, [
    ("In scope (v1)", "One vertical (steel); daily horizon; decoupled forecast + event + "
                      "attribution pipeline; offline backtest + case-study notebook."),
    ("Out of scope (v1)", "Multi-commodity fan-out; end-to-end jointly-trained multimodal model; "
                          "live production alerting infra; intraday/tick data; trade execution."),
    ("Planned v2", "Moirai-2 with macro covariates; LLM explanation layer; additional verticals "
                   "(oil, semiconductors, ag); streaming alerts to Slack/Teams."),
])
r += 1
r = section(ws, r, "Assumptions & Constraints")
r = kv_rows(ws, r, [
    ("Assumptions", "Public data is sufficient for v1; Federal Register API is the authoritative, "
                    "timestamped event source; daily cadence is adequate lead time."),
    ("Constraints", "Free / low-cost data only; runs on CPU or a modest single GPU; interpretable, "
                    "auditable components preferred over black-box end-to-end training."),
    ("Success criteria", "Beats seasonal-naive & ARIMA on interval coverage; positive median "
                         "lead-time on labeled shock events; >=60% top-3 attribution accuracy."),
])

# ----------------------------------------------------------------------------
# SHEET 4 - USE CASE DIAGRAM
# ----------------------------------------------------------------------------
ws = wb.create_sheet("4. Use Case Diagram")
ws.sheet_view.showGridLines = False
# fine grid
for col in range(1, 34):
    ws.column_dimensions[get_column_letter(col)].width = 3.1
for rr in range(1, 60):
    ws.row_dimensions[rr].height = 15

# title spanning
ws.merge_cells("A1:AG1")
c = ws["A1"]; c.value = "Use Case Diagram"; c.font = F(16, True, WHITE); c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 28
ws.merge_cells("A2:AG2")
c = ws["A2"]; c.value = "Actors, system use cases, and relationships (UML-style)"
c.font = F(10, False, WHITE, italic=True); c.fill = fill(BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# System boundary drawn as a border rectangle (NOT merged, so inner cells stay free)
def boundary(ws, r1, c1, r2, c2, fillcol, edge):
    sd = Side(style="medium", color=edge)
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = fill(fillcol)
            left = sd if cc == c1 else None
            right = sd if cc == c2 else None
            top = sd if rr == r1 else None
            bottom = sd if rr == r2 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

boundary(ws, 5, 11, 40, 24, BOUND, STEEL)
ws.merge_cells(start_row=5, start_column=11, end_row=5, end_column=24)
lab = ws.cell(row=5, column=11, value="  << system >>  Tariff & Geopolitical Shock EWS")
lab.font = F(10, True, STEEL); lab.alignment = Alignment(horizontal="left", vertical="center")

# Left actors
def actor(ws, r, c1, c2, name):
    box(ws, r, c1, r + 2, c2, "«actor»\n" + name, CARD_ACT, NAVY, bold=True, size=9)

actor(ws, 7, 2, 9, "Commodity\nRisk Analyst")
actor(ws, 14, 2, 9, "Supply-Chain\nRisk Manager")
actor(ws, 21, 2, 9, "Trade-Credit\nUnderwriter")
actor(ws, 34, 2, 9, "Scheduler /\nSystem (cron)")

# Right actors (external systems)
actor(ws, 9, 26, 33, "Market/Macro\nData Providers")
actor(ws, 20, 26, 33, "Federal Register /\nNews Event Feeds")

# Use cases (ovals-as-boxes) inside boundary, cols 12-23
def uc(ws, r, text, fillc=CARD_UC):
    box(ws, r, 12, r + 1, 23, text, fillc, NAVY, bold=False, size=9,
        border=Border(left=thin, right=thin, top=thin, bottom=thin))

uc(ws, 7,  "UC1  Ingest price & macro series")
uc(ws, 10, "UC2  Ingest policy / geo events")
uc(ws, 13, "UC3  Generate calibrated forecast")
uc(ws, 16, "UC4  Monitor for interval breach")
uc(ws, 19, "UC5  Detect changepoint (cross-check)")
uc(ws, 22, "UC6  Attribute break to event")
uc(ws, 25, "UC7  Emit early-warning alert")
uc(ws, 28, "UC8  View backtest / case-study report")
uc(ws, 31, "UC9  Configure vertical & thresholds")

# connector hints (arrows) - simple text arrows between actor edge and UC edge
def arrow(ws, r, c, ch="—▶", color=GREY):
    cell = ws.cell(row=r, column=c, value=ch)
    cell.font = F(9, True, color); cell.alignment = CENTER

for r in (8, 15, 22):
    arrow(ws, r, 10)      # primary actors -> use cases
arrow(ws, 35, 10)         # scheduler -> use cases
arrow(ws, 10, 25, "◀—")   # Market/Macro Data Providers (rows 9-11) serve UC1/UC2
arrow(ws, 21, 25, "◀—")   # Federal Register / News feeds (rows 20-22) serve UC2

# Relationship legend / catalog note below
box(ws, 42, 2, 42, 24, "Relationships", BLUE, WHITE, size=11)
rels = [
    "UC3 «include» UC1        (forecast requires ingested series)",
    "UC4 «include» UC3        (breach monitoring runs on the calibrated forecast)",
    "UC6 «include» UC2, UC4   (attribution needs events + a detected breach)",
    "UC6 «extend»  UC5        (changepoint agreement raises attribution confidence)",
    "UC7 «include» UC6        (alert carries the attributed driver)",
    "Primary actors: Risk Analyst, Supply-Chain Manager, Underwriter (consume UC7/UC8)",
    "Secondary actors: Scheduler triggers UC1-UC7; Data/Event feeds serve UC1/UC2",
]
for i, t in enumerate(rels):
    er = 43 + i
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=24)
    cell = ws.cell(row=er, column=2, value="•  " + t)
    cell.font = F(9.5); cell.alignment = LEFT
    if i % 2 == 1:
        for cc in range(2, 25):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 4b -> merged into same workbook: Use Case Catalog on its own sheet
# ----------------------------------------------------------------------------
ws = wb.create_sheet("4b. Use Case Catalog")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 8, "B": 26, "C": 20, "D": 40, "E": 22})
title_block(ws, "Use Case Catalog", "Detailed description of each use case in the diagram", "E")
table_header(ws, 4, ["ID", "Use Case", "Primary Actor", "Description / Main Flow", "Trigger"])
uc_rows = [
    ["UC1", "Ingest price & macro series", "Scheduler", "Pull HRC/SLX prices + FRED macro, normalize to daily parquet.", "Daily schedule"],
    ["UC2", "Ingest policy / geo events", "Scheduler", "Fetch Federal Register + news items, normalize into event table with timestamps.", "Daily schedule"],
    ["UC3", "Generate calibrated forecast", "Risk Analyst", "TimesFM 2.5 zero-shot forecast, calibrated online with ACI for intervals.", "On new data"],
    ["UC4", "Monitor for interval breach", "System", "Compare actuals to the calibrated interval; flag breaches.", "On new actuals"],
    ["UC5", "Detect changepoint", "System", "PELT / BOCPD on residuals as an independent structural-break check.", "On new actuals"],
    ["UC6", "Attribute break to event", "Risk Analyst", "Rank recent events by relevance x severity x recency; pick top driver.", "On breach"],
    ["UC7", "Emit early-warning alert", "Risk Analyst", "Produce alert: series, breach, attributed event, source link, lead time.", "On breach"],
    ["UC8", "View backtest / case study", "Analyst / Reviewer", "Inspect coverage, MASE, lead-time, attribution accuracy, dated examples.", "On demand"],
    ["UC9", "Configure vertical & thresholds", "Analyst", "Edit YAML: series IDs, HS codes, keywords, horizons, breach thresholds.", "On demand"],
]
table_rows(ws, 5, uc_rows)

# ----------------------------------------------------------------------------
# SHEET 5 - DATA SOURCES & ACCESS
# ----------------------------------------------------------------------------
ws = wb.create_sheet("5. Data Sources & Access")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 18, "B": 26, "C": 26, "D": 12, "E": 12, "F": 30})
title_block(ws, "Data Sources & Access", "What feeds the system, how it is accessed, and cadence", "F")
table_header(ws, 4, ["Layer", "Source", "Access Method", "Cost", "Cadence", "Notes / Fields"])
data_rows = [
    ["Target series", "Steel: HRC futures / SLX ETF proxy", "yfinance (Python, no key)", "Free", "Daily", "Close/OHLCV; SLX is the clean liquid proxy for v1."],
    ["Macro / exogenous", "FRED: steel PPI, industrial production, import price index, USD index", "fredapi + free API key", "Free", "Daily / Monthly", "Covariates for v2 (Moirai); context for v1."],
    ["Trade volumes", "US Census international trade (steel HS 72xx)", "Census API + free key", "Free", "Monthly", "Import/export volume & value by HS code."],
    ["Shipping cost", "Freightos Baltic Index (FBX) / Baltic Dry proxy", "Free tier / scrape", "Free", "Daily / Weekly", "Container & dry-bulk freight as shock transmission channel."],
    ["Event stream (core)", "Federal Register (Section 232/301, USTR, Commerce)", "Federal Register API (JSON)", "Free", "Daily", "AUTHORITATIVE, timestamped, searchable by agency+keyword; makes attribution verifiable."],
    ["Event stream (announcement-dated)", "Daily TPU index (Caldara-Iacoviello, JME 2020)", "XLSX download, no key", "Free", "Daily (updated ~monthly)", "Expert newspaper-count index; spikes are announcement-dated. Primary news layer."],
    ["Event stream (supplement)", "GDELT news-volume spikes (EPU-style)", "GDELT DOC 2.0 API", "Free", "Daily", "Fallback: API hard-throttles multi-year pulls; incremental cache implemented."],
]
data_fills = {4: MINT}  # highlight the core event source
table_rows(ws, 5, data_rows, fills=data_fills)

r = 5 + len(data_rows) + 1
box(ws, r, 1, r, 6, "Design decisions & data governance", BLUE, WHITE, size=11)
notes = [
    "The Federal Register API is the anchor of the whole project: free, structured JSON, timestamped, and queryable by agency + date + keyword. It makes 'which event caused the break' auditable rather than hand-wavy.",
    "All sources are free / public. No proprietary or licensed market feeds are required for v1, so the repo is fully reproducible by a reviewer.",
    "Raw pulls are cached to gitignored parquet; only derived, shareable artifacts are committed.",
    "A frozen holdout window (most recent ~6 months) is never touched during development to keep the backtest honest.",
    "Rate limits: batch pulls, cache aggressively, and stamp each record with a pull timestamp for reproducibility.",
]
for i, t in enumerate(notes):
    er = r + 1 + i
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=6)
    cell = ws.cell(row=er, column=1, value="•  " + t)
    cell.font = F(9.5); cell.alignment = LEFT_T; cell.border = BORDER_THIN
    ws.row_dimensions[er].height = max(26, 14 * (1 + len(t)//110))
    if i % 2 == 1:
        for cc in range(1, 7):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 6 - MODEL SELECTION & RATIONALE
# ----------------------------------------------------------------------------
ws = wb.create_sheet("6. Model Selection")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 20, "B": 12, "C": 26, "D": 26, "E": 26, "F": 12})
title_block(ws, "Model Selection & Rationale", "Trade study of time-series foundation models and supporting components", "F")

# Trade study
ws.cell(row=4, column=1, value="Forecasting backbone - trade study").font = F(11, True, NAVY)
table_header(ws, 5, ["Criterion", "Weight", "TimesFM 2.5 (chosen)", "Chronos-2", "Moirai-2", "Winner"])
trade = [
    ["Native probabilistic output", "High", "Yes - continuous quantile head (deciles)", "Yes - quantile forecasts", "Yes - distributional", "Tie"],
    ["Tooling / preflight for local run", "High", "Vetted skill + system preflight checker", "Clean (AutoGluon)", "Good, newer", "TimesFM 2.5"],
    ["Covariate / multivariate fusion", "Med", "Supported via XReg (v2)", "Limited", "Strong (best-in-class)", "Moirai-2"],
    ["Compute footprint", "High", "200M, ~1.5 GB RAM, CPU-friendly", "CPU / modest GPU", "Modest-heavy GPU", "TimesFM 2.5"],
    ["Context length", "Med", "Up to 16,384 points", "Moderate", "Moderate", "TimesFM 2.5"],
    ["Fit for v1 goals", "High", "Best: calibrated zero-shot, CPU, ready tooling", "Strong alternative", "Best for v2 covariates", "TimesFM 2.5"],
]
tfills = {}
table_rows(ws, 6, trade, zebra=True)
# tint the chosen column (col C = 3)
for r in range(6, 6 + len(trade)):
    ws.cell(row=r, column=3).fill = fill(MINT)
    ws.cell(row=r, column=6).font = F(9.5, True, MINT_DK)

r = 6 + len(trade) + 1
box(ws, r, 1, r, 6, "Decision", MINT_DK, WHITE, size=11)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=6)
dec = ("DECISION: Use TimesFM 2.5 (200M, PyTorch) as the v1 forecasting backbone.\n"
       "WHY: Its continuous quantile head returns calibrated decile intervals zero-shot, it is CPU-friendly "
       "(~1.5 GB RAM), and it ships a vetted skill plus a mandatory preflight checker for safe local inference. "
       "The raw q10/q90 band is then widened to the 90% target with split-conformal CQR (Romano 2019). "
       "Chronos-2 remains a drop-in alternative (kept in src/forecast/chronos_model.py); Moirai-2 stays the "
       "planned v2 upgrade for covariate fusion. This supersedes the v1.0 choice of Chronos-2: TimesFM 2.5's "
       "quantile head removes the earlier 'point-centric' objection.")
cell = ws.cell(row=r, column=1, value=dec)
cell.font = F(9.5); cell.alignment = LEFT_T; cell.fill = fill(MINT)
cell.border = BORDER_THIN

r += 5
ws.cell(row=r, column=1, value="Supporting components").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Component", "Choice", "Purpose", "Why this choice", "", ""])
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
comps = [
    ["Interval calibration", "Split-conformal CQR (Romano 2019)", "Widen the q10/q90 band to guaranteed target coverage", "Makes the forecast a trustworthy risk signal, not just a line; defensible metric."],
    ["Event embedding", "sentence-transformers (MiniLM / bge-small)", "Score event relevance to the commodity", "Local, free, fast; no external LLM dependency in v1."],
    ["Severity tagging", "Rule layer (tariff %, 'Section 232', 'effective immediately')", "Grade event impact", "Transparent and auditable; beats a black box for credibility."],
    ["Changepoint detection", "ruptures (PELT) or BOCPD", "Independent structural-break check on residuals", "Cross-checks interval breaches; agreement raises confidence."],
    ["Attribution", "relevance x severity x recency ranking", "Name the likely driving event", "Simple, explainable, and tied to a verifiable source link."],
    ["Baselines", "Seasonal-naive, ARIMA, GARCH", "Benchmark gauntlet", "Must be beaten on calibration or explained; reviewers look for this."],
]
for i, row in enumerate(comps):
    er = r + 1 + i
    for j in range(4):
        cell = ws.cell(row=er, column=1 + j, value=row[j])
        cell.font = F(9.5, bold=(j == 0)); cell.alignment = LEFT_T; cell.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=4, end_row=er, end_column=6)
    for cc in range(1, 7):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 7):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)
    ws.row_dimensions[er].height = 30

# ----------------------------------------------------------------------------
# SHEET 7 - TEXT-FUSION ARCHITECTURE
# ----------------------------------------------------------------------------
ws = wb.create_sheet("7. Text-Fusion Arch")
ws.sheet_view.showGridLines = False
for col in range(1, 40):
    ws.column_dimensions[get_column_letter(col)].width = 2.9
for rr in range(1, 60):
    ws.row_dimensions[rr].height = 15

ws.merge_cells("A1:AM1")
c = ws["A1"]; c.value = "Text-Fusion Architecture"; c.font = F(16, True, WHITE); c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 28
ws.merge_cells("A2:AM2")
c = ws["A2"]; c.value = "Decoupled numeric + text pipeline (interpretable, low-compute). Data flows left to right."
c.font = F(10, False, WHITE, italic=True); c.fill = fill(BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Column swimlane labels
box(ws, 4, 2, 4, 9,  "INPUTS", GREY, WHITE, size=9)
box(ws, 4, 11, 4, 20, "ENCODE / MODEL", GREY, WHITE, size=9)
box(ws, 4, 22, 4, 30, "FUSE / DETECT", GREY, WHITE, size=9)
box(ws, 4, 32, 4, 38, "OUTPUT", GREY, WHITE, size=9)

# NUMERIC path (top)
box(ws, 6, 2, 8, 9, "Numeric series\n(HRC/SLX, macro, FBX)", LIGHT, NAVY, size=9)
box(ws, 6, 11, 8, 20, "TimesFM 2.5 zero-shot\nquantile forecast", CARD_ACT, NAVY, size=9)
box(ws, 10, 11, 12, 20, "ACI online calibration\n(CQR offline)", CARD_ACT, NAVY, size=9)

# TEXT path (bottom)
box(ws, 14, 2, 16, 9, "Event stream\n(Federal Register, news)", LIGHT, NAVY, size=9)
box(ws, 14, 11, 16, 20, "Sentence-transformer\nembeddings", AMBER, AMBER_DK, size=9)
box(ws, 18, 11, 20, 20, "Rule-based severity\ntagging", AMBER, AMBER_DK, size=9)

# Fuse / detect column
box(ws, 6, 22, 8, 30, "Interval-breach\nmonitor  (UC4)", MINT, MINT_DK, size=9)
box(ws, 10, 22, 12, 30, "Changepoint check\nPELT / BOCPD (UC5)", MINT, MINT_DK, size=9)
box(ws, 14, 22, 16, 30, "Event relevance +\ntiming scorer", MINT, MINT_DK, size=9)
box(ws, 18, 22, 20, 30, "Attribution engine\nrel × sev × recency", NAVY, WHITE, size=9)

# Output
box(ws, 9, 32, 13, 38, "EARLY-WARNING\nALERT\n(series, breach,\ndriver + source link,\nlead time)", STEEL, WHITE, size=9)
box(ws, 16, 32, 19, 38, "Backtest / case-\nstudy report\n(coverage, lead time,\nattribution acc.)", BLUE, WHITE, size=9)

# arrows
def har(ws, r, c):
    cell = ws.cell(row=r, column=c, value="▶"); cell.font = F(11, True, GREY); cell.alignment = CENTER
def var(ws, r, c):
    cell = ws.cell(row=r, column=c, value="▼"); cell.font = F(11, True, GREY); cell.alignment = CENTER

har(ws, 7, 10); har(ws, 15, 10)          # inputs -> encode
var(ws, 9, 15)                            # chronos -> conformal
var(ws, 17, 15)                           # embed -> severity
har(ws, 11, 21)                           # conformal -> breach monitor
har(ws, 7, 31); har(ws, 11, 31)           # detect -> alert (up)
har(ws, 19, 31)                           # attribution -> report
# feed lines into attribution
var(ws, 13, 26); var(ws, 17, 26)

# Legend
box(ws, 23, 2, 23, 20, "Legend", BLUE, WHITE, size=10)
legend = [
    ("Numeric path", LIGHT), ("Model / encoder", CARD_ACT), ("Text path", AMBER),
    ("Fuse / detect", MINT), ("Output artifact", STEEL),
]
lc = 2
for name, col in legend:
    box(ws, 24, lc, 24, lc + 2, "", col)
    cell = ws.cell(row=24, column=lc + 3, value=name); cell.font = F(9); cell.alignment = LEFT
    lc += 7

# Design principle notes
box(ws, 26, 2, 26, 38, "Why decoupled (not end-to-end multimodal) in v1", BLUE, WHITE, size=10)
princ = [
    "Interpretability: every hop (forecast, breach, event score, attribution) is inspectable and auditable, which is the credibility of the product.",
    "Compute: no joint multimodal training; runs on CPU / a modest GPU with free public data.",
    "Data reality: not enough labeled joint data to train an end-to-end model well; a retrieval + scoring design generalizes better at this stage.",
    "Upgrade path: v2 can add a Moirai-2 covariate backbone and an LLM explanation layer over the same scaffolding without redesign.",
]
for i, t in enumerate(princ):
    er = 27 + i
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=38)
    cell = ws.cell(row=er, column=2, value="•  " + t); cell.font = F(9.5); cell.alignment = LEFT_T
    if i % 2 == 1:
        for cc in range(2, 39):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 7b - Component catalog for architecture
# ----------------------------------------------------------------------------
ws = wb.create_sheet("7b. Component Catalog")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 24, "B": 18, "C": 30, "D": 30, "E": 20})
title_block(ws, "Architecture Component Catalog", "Each block in the data flow, its tech and contract", "E")
table_header(ws, 4, ["Component", "Tech", "Input", "Output", "Owner module"])
carch = [
    ["Series loaders", "yfinance / fredapi / Census", "API pulls", "Normalized daily parquet", "src/data/*"],
    ["Event loader", "Federal Register API + news", "Keyword+date queries", "Event table (ts, text, agency)", "src/data/events.py"],
    ["Forecaster", "TimesFM 2.5 (Chronos-2 alt)", "Series history", "Quantile forecast (deciles)", "src/forecast/timesfm_model.py"],
    ["Calibrator (live)", "ACI online (CQR for offline studies)", "Raw q10/q90 band history", "Calibrated 90% interval, regime-robust", "src/forecast/conformal.py"],
    ["GARCH cross-check", "GARCH(1,1) band (arch)", "Price history", "Independent volatility band, garch_agrees flag", "src/eval/backtest.py"],
    ["Band history cache", "Parquet incremental cache", "New H-day blocks only", "data/bands_<vertical>.parquet", "src/pipeline.py"],
    ["Breach monitor", "NumPy logic", "Actuals + interval", "Breach flags", "src/pipeline.py"],
    ["Changepoint", "ruptures / BOCPD", "Forecast residuals", "Changepoint indices", "src/detect/changepoint.py"],
    ["Event embedder", "sentence-transformers", "Event text", "Embeddings + relevance", "src/events/embed.py"],
    ["Severity tagger", "Rule engine", "Event text", "Severity score", "src/events/severity.py"],
    ["Attribution", "Ranking", "Events + breach + changepoint", "Top driver + link", "src/events/attribute.py"],
    ["Evaluator", "pandas / metrics", "Forecasts, events, labels", "Coverage, MASE, lead time, attr acc", "src/eval/*"],
]
table_rows(ws, 5, carch)

# ----------------------------------------------------------------------------
# SHEET 8 - FILE / REPO STRUCTURE
# ----------------------------------------------------------------------------
ws = wb.create_sheet("8. File Structure")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 52, "B": 54})
title_block(ws, "Repository / File Structure", "Directory layout with responsibility of each path", "B")

tree = [
    ("tariff-shock-ews/", "Repository root", True),
    ("├─ README.md", "Problem, one dated backtested example up top, limitations", False),
    ("├─ pyproject.toml", "deps: timesfm[torch], arch, ruptures, sentence-transformers (chronos optional)", False),
    ("├─ config/", "Per-vertical configuration", True),
    ("│   ├─ steel.yaml", "Series IDs, HS codes, keywords, horizons, thresholds", False),
    ("│   └─ aluminum.yaml", "Second vertical (ALI=F), zero code changes (v2-W2)", False),
    ("├─ src/", "Application source", True),
    ("│   ├─ data/", "Ingestion layer", True),
    ("│   │   ├─ prices.py", "yfinance / FRED loaders -> parquet", False),
    ("│   │   ├─ trade.py", "US Census trade (HS 72xx)", False),
    ("│   │   ├─ events.py", "Federal Register + news -> normalized event table", False),
    ("│   │   ├─ news_tpu.py", "Daily TPU index (announcement-dated events, primary)", False),
    ("│   │   └─ news_gdelt.py", "GDELT volume spikes (fallback; API throttles)", False),
    ("│   ├─ forecast/", "Forecasting layer", True),
    ("│   │   ├─ timesfm_model.py", "TimesFM 2.5 zero-shot forecast + quantiles (v1 default)", False),
    ("│   │   ├─ timesfm_xreg.py", "Covariate (XReg) forecaster; evaluated, does not ship (v2-W1)", False),
    ("│   │   ├─ chronos_model.py", "Chronos-2 forecaster (alternative backbone)", False),
    ("│   │   └─ conformal.py", "Split-conformal CQR + ACI calibration", False),
    ("│   ├─ events/", "Event reasoning layer", True),
    ("│   │   ├─ embed.py", "Sentence-transformer embeddings", False),
    ("│   │   ├─ severity.py", "Rule-based severity tagging", False),
    ("│   │   └─ attribute.py", "relevance x severity x recency ranking", False),
    ("│   ├─ detect/", "Structural-break layer", True),
    ("│   │   └─ changepoint.py", "ruptures / BOCPD on residuals", False),
    ("│   ├─ pipeline.py", "Orchestrates forecast -> monitor -> attribute -> alert", False),
    ("│   └─ eval/", "Evaluation layer", True),
    ("│       ├─ backtest.py", "Rolling-origin; coverage, MASE", False),
    ("│       └─ event_eval.py", "Lead-time, precision/recall, attribution accuracy", False),
    ("├─ notebooks/", "Analysis & demo", True),
    ("│   └─ 01_case_study_steel.ipynb", "Interactive walkthrough of the production path", False),
    ("├─ tools/", "Repo meta-tooling", True),
    ("│   └─ build_design_doc.py", "Regenerates docs/ solution design workbook", False),
    ("├─ scripts/", "Runnable build-step scripts", True),
    ("│   ├─ step1_forecast_slx.py", "Step 1: SLX TimesFM forecast + CQR interval + baselines", False),
    ("│   ├─ step2_events_attribution.py", "Step 2: breach detection + Federal Register attribution", False),
    ("│   ├─ step3_close_gaps.py", "Step 3: ACI, GARCH baseline, frozen holdout, permutation test", False),
    ("│   ├─ step4_hrc_attribution.py", "Step 4: pre-registered attribution retest on HRC futures", False),
    ("│   ├─ step5_forward_eval.py", "Step 5: forward early-warning eval (G3), ACI flags vs shocks", False),
    ("│   ├─ step6_announcement_attribution.py", "Step 6: announcement-dated (TPU) attribution retest", False),
    ("│   ├─ step7_alert_precision.py", "Step 7: pre-specified alert-precision filter evaluation", False),
    ("│   └─ step8_covariates_eval.py", "Step 8 (v2-W1): XReg vs v1 pre-registered eval", False),
    ("├─ data/", "Gitignored raw + parquet cache", True),
    ("└─ outputs/", "Generated artifacts", True),
    ("    ├─ figures/", "Interval-breach plots for the writeup", False),
    ("    └─ alerts/", "Machine-readable alert artifacts per vertical/date (v2-W3)", False),
]
table_header(ws, 4, ["Path", "Responsibility"])
for i, (path, resp, isdir) in enumerate(tree):
    er = 5 + i
    c1 = ws.cell(row=er, column=1, value=path)
    c1.font = F(9.5, bold=isdir, color=(NAVY if isdir else "000000"), name="Consolas")
    c1.alignment = LEFT; c1.border = BORDER_THIN
    c2 = ws.cell(row=er, column=2, value=resp)
    c2.font = F(9.5); c2.alignment = LEFT; c2.border = BORDER_THIN
    if isdir:
        c1.fill = fill(LIGHT); c2.fill = fill(LIGHT)
    elif i % 2 == 1:
        c1.fill = fill(LIGHTER); c2.fill = fill(LIGHTER)

r = 5 + len(tree) + 1
box(ws, r, 1, r, 2, "Recommended build order (always keep something working)", BLUE, WHITE, size=11)
order = [
    "1. data/prices.py + forecast/chronos_model.py  ->  a zero-shot forecast plotted (Day 1-2)",
    "2. forecast/conformal.py + eval/backtest.py     ->  calibrated intervals + coverage (Day 3-4)",
    "3. data/events.py + events/*                     ->  event table + attribution (Day 5-7)",
    "4. eval/event_eval.py on real 2025-26 events     ->  the money chart (Day 8-9)",
    "5. notebooks/01 + README case study             ->  the shareable artifact (Day 10)",
]
for i, t in enumerate(order):
    er = r + 1 + i
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=2)
    cell = ws.cell(row=er, column=1, value=t)
    cell.font = F(9.5, name="Consolas"); cell.alignment = LEFT; cell.border = BORDER_THIN
    if i % 2 == 1:
        for cc in (1, 2):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 9 - EVALUATION & BACKTEST DESIGN
# ----------------------------------------------------------------------------
ws = wb.create_sheet("9. Evaluation & Backtest")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 26, "B": 30, "C": 30, "D": 26})
title_block(ws, "Evaluation & Backtest Design", "How every claim is made dated, falsifiable, and defensible", "D")

table_header(ws, 4, ["Evaluation", "Method", "Metric(s)", "Pass bar"])
evals = [
    ["Forecast quality", "Rolling-origin (expanding window) backtest", "MASE (not raw RMSE)", "Beat seasonal-naive"],
    ["Interval calibration (headline)", "Empirical coverage vs nominal", "Coverage %, mean interval width", "90% interval covers ~90% (+/- tol)"],
    ["Baseline gauntlet", "Compare vs classical models", "MASE + coverage vs baselines", "Beat, or explain, seasonal-naive/ARIMA/GARCH"],
    ["Event early-warning (money metric)", "Labeled 2025-26 tariff events from Federal Register", "Lead time (days), precision/recall on flagged shocks", "Positive median lead time"],
    ["Attribution accuracy", "Top-ranked event vs ground-truth cause", "Top-1 / Top-3 accuracy", ">= 60-70% Top-3"],
    ["Robustness", "Frozen ~6-month holdout never seen in dev", "Metrics hold on holdout", "No large degradation"],
]
efills = {1: MINT, 3: MINT}
table_rows(ws, 5, evals, fills=efills)

r = 5 + len(evals) + 1
box(ws, r, 1, r, 4, "Evaluation principles", BLUE, WHITE, size=11)
pr = [
    "Coverage is the headline, not accuracy: a foundation model that cannot beat seasonal-naive on calibration is a red flag reviewers look for.",
    "The event early-warning eval is the demo: 'here is the interval on date X, the Federal Register tariff notice on date Y, the price move, and the lead time.'",
    "Every reported claim is dated and links to a verifiable source, so it is falsifiable rather than a vague accuracy number.",
    "A clean holdout is frozen before development and never touched, to keep the backtest honest.",
]
for i, t in enumerate(pr):
    er = r + 1 + i
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=4)
    cell = ws.cell(row=er, column=1, value="•  " + t)
    cell.font = F(9.5); cell.alignment = LEFT_T; cell.border = BORDER_THIN
    ws.row_dimensions[er].height = max(26, 14*(1+len(t)//120))
    if i % 2 == 1:
        for cc in range(1, 5):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 10 - NFRs & RISKS
# ----------------------------------------------------------------------------
ws = wb.create_sheet("10. NFRs & Risks")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 22, "B": 44, "C": 14, "D": 34})
title_block(ws, "Non-Functional Requirements & Risks", "Quality attributes and how key risks are mitigated", "D")

ws.cell(row=4, column=1, value="Non-Functional Requirements").font = F(11, True, NAVY)
table_header(ws, 5, ["Attribute", "Requirement", "", ""])
ws.merge_cells("C5:D5")
nfr = [
    ["Reproducibility", "Fully reproducible from free public data; pinned deps; cached pulls with timestamps."],
    ["Interpretability", "Every stage inspectable; no black-box end-to-end model in v1."],
    ["Cost", "$0 data; runs on CPU or a single modest GPU."],
    ["Performance", "Daily batch; full pipeline for one vertical runs in minutes, not hours."],
    ["Auditability", "Attribution always cites a timestamped Federal Register source link."],
    ["Extensibility", "New vertical = new YAML; v2 backbone swap without redesign."],
]
for i, (a, req) in enumerate(nfr):
    er = 6 + i
    c1 = ws.cell(row=er, column=1, value=a); c1.font = F(9.5, True, NAVY); c1.alignment = LEFT_T; c1.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=4)
    c2 = ws.cell(row=er, column=2, value=req); c2.font = F(9.5); c2.alignment = LEFT_T
    for cc in range(1, 5):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 5):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

r = 6 + len(nfr) + 1
ws.cell(row=r, column=1, value="Risk Register").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Risk", "Description", "Severity", "Mitigation"])
risks = [
    ["Spurious attribution", "Model blames the wrong event for a break.", "High", "Require changepoint agreement; report Top-3; always show source link for human check."],
    ["Poor calibration", "Intervals over/under-cover on real data.", "High", "Conformal wrapper; report coverage explicitly; fall back to wider quantiles."],
    ["Event data noise", "News feeds are noisy / low signal.", "Med", "Anchor on Federal Register; use news only to supplement; severity rules."],
    ["Overfitting to known events", "Tuning to make past shocks look predictable.", "Med", "Frozen holdout; rolling-origin; no peeking at holdout in dev."],
    ["Scope creep", "Multi-commodity fan-out too early.", "Med", "v1 frozen to steel; fan-out gated behind a working single vertical."],
    ["Data availability", "Free source changes / rate limits.", "Low", "Cache aggressively; proxies (SLX for HRC); documented fallbacks."],
]
sevfill = {"High": AMBER, "Med": LIGHTER, "Low": GREY_LT}
for i, row in enumerate(risks):
    er = r + 1 + i
    for j, val in enumerate(row):
        c = ws.cell(row=er, column=1 + j, value=val)
        c.font = F(9.5, bold=(j == 0)); c.alignment = LEFT_T; c.border = BORDER_THIN
    ws.cell(row=er, column=3).fill = fill(sevfill.get(row[2], WHITE))
    ws.cell(row=er, column=3).alignment = CENTER
    ws.row_dimensions[er].height = 30

# ----------------------------------------------------------------------------
# SHEET 11 - BUILD LOG & VALIDATION
# ----------------------------------------------------------------------------
ws = wb.create_sheet("11. Build Log & Validation")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 30, "B": 30, "C": 26, "D": 16, "E": 18})
title_block(ws, "Build Log & Validation", "What has actually been built and measured, versus the documented goals", "E")

STAT = {"Met": MINT, "Partial": AMBER, "Deferred": AMBER, "Next": LIGHTER, "Open": AMBER}

r = 4
ws.cell(row=r, column=1, value="Build progression (against the file-structure build order)").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Build order step", "Deliverable", "Status", "Date", "Evidence"])
prog = [
    ["1. prices + forecaster", "Zero-shot forecast plotted", "Met", "2026-07-09", "outputs/figures/*.png"],
    ["2. conformal + backtest", "Calibrated intervals + coverage", "Met", "2026-07-09", "step1_summary.json"],
    ["3. events + attribution", "Event table + breach attribution", "Met", "2026-07-09", "step2_summary.json"],
    ["4. event eval", "Lead-time / precision / attribution acc.", "Met", "2026-07-09", "step5_forward_eval.json"],
    ["5. notebook + writeup", "Shareable case study", "Partial", "2026-07-09", "README worked examples"],
]
pf = {i: STAT.get(row[2]) for i, row in enumerate(prog)}
table_rows(ws, r + 1, prog, fills=pf)

r = r + 1 + len(prog) + 2
ws.cell(row=r, column=1, value="Step-1 goal verification (forecast + calibration)").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Goal / eval criterion", "Target / pass bar", "Measured result", "Status", "Notes"])
verif = [
    ["G1 Calibrated forecast", "90% interval covers ~90%", "91.5% coverage", "Met", "CQR offset Q=1.55 USD"],
    ["Interval calibration (headline)", "coverage ~ nominal", "raw 80% -> 79.2%; cal -> 91.5%", "Met", "TimesFM deciles near-calibrated"],
    ["Forecast quality (MASE)", "beat seasonal-naive", "TimesFM 4.33 vs naive 4.21", "Partial", "Explained: efficient ETF ~ random walk"],
    ["Baseline gauntlet", "naive / ARIMA / GARCH", "naive + ARIMA + GARCH(1,1)", "Met", "GARCH added (step3), 87.5% holdout cov"],
    ["Robustness", "frozen ~6-month holdout", "frozen 2026 holdout + ACI", "Met", "fixed 80% fwd -> ACI holds ~90%"],
    ["G4 Checkable", "dated, falsifiable, reproducible", "dated example + script", "Met", "step1_forecast_slx.py"],
]
vf = {i: STAT.get(row[3]) for i, row in enumerate(verif)}
table_rows(ws, r + 1, verif, fills=vf)

r = r + 1 + len(verif) + 2
ws.cell(row=r, column=1, value="Step-1 measured results (SLX, 2016-01-04 to 2026-07-08)").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Metric", "Value", "", "", ""])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
res = [
    ["Backbone / calibration", "TimesFM 2.5 zero-shot + split-conformal CQR"],
    ["Horizon", "10 business days"],
    ["Raw q10-q90 coverage (nominal 80%)", "79.2%"],
    ["CQR-calibrated coverage (target 90%)", "91.5%"],
    ["CQR offset Q", "1.55 (USD)"],
    ["MASE  TimesFM / ARIMA / naive", "4.33 / 4.22 / 4.21"],
    ["Headline holdout (2026-06-24 -> 07-08)", "0 interval breaches"],
]
for i, (k, v) in enumerate(res):
    er = r + 1 + i
    c1 = ws.cell(row=er, column=1, value=k); c1.font = F(9.5, True, NAVY); c1.alignment = LEFT_T; c1.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=5)
    c2 = ws.cell(row=er, column=2, value=v); c2.font = F(9.5); c2.alignment = LEFT_T
    for cc in range(1, 6):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 6):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

r = r + 1 + len(res) + 2
ws.cell(row=r, column=1, value="Step-2 measured results (event attribution, retrospective)").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Metric", "Value", "", "", ""])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
res2 = [
    ["Events (Federal Register proclamations + rules)", "255 (2018-02 to 2026-06)"],
    ["Full-sample CQR realized coverage", "90.1% (target 90%, Q=0.73)"],
    ["Breach-days flagged", "209 / 2110 (~10% tail)"],
    ["Volatility-regime changepoints", "18 (2020 crash, 2022, 2025 turmoil)"],
    ["Breaches w/ tariff event in 14d vs base rate", "72% vs 66% (density-driven, honest)"],
    ["Dated example", "2022-06-13 breach -> 2022-06-03 steel proclamation (rel 0.64)"],
]
for i, (k, v) in enumerate(res2):
    er = r + 1 + i
    c1 = ws.cell(row=er, column=1, value=k); c1.font = F(9.5, True, NAVY); c1.alignment = LEFT_T; c1.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=5)
    c2 = ws.cell(row=er, column=2, value=v); c2.font = F(9.5); c2.alignment = LEFT_T
    for cc in range(1, 6):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 6):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

r = r + 1 + len(res2) + 2
ws.cell(row=r, column=1, value="Gap closure (literature-backed methods)").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Gap", "Method (source)", "Result", "Status", ""])
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
closed = [
    ["A. Frozen holdout", "Untouched trailing 6-mo (2026)", "Fixed CQR 80% fwd coverage (under-covers)", "Closed"],
    ["B. Regime-robust calibration", "Adaptive Conformal Inference (Gibbs & Candes 2021)", "Holds ~90%, far more stable than fixed offset", "Closed"],
    ["C. GARCH baseline", "GARCH(1,1) volatility band (arch)", "87.5% holdout coverage (beat fixed offset fwd)", "Closed"],
    ["D. Attribution significance", "Event-study circular-shift permutation test", "72.2% vs 66.0% base, p=0.134 -> not significant", "Tested"],
    ["E. Point vs naive", "TSFM benchmarks (arxiv 2601.06371, cloud FM study)", "Expected on efficient daily prices; not a bug", "Closed"],
    ["F. Attribution retest (HRC)", "Pre-registered primary: HRC=F, sev>=0.7 events, 5d window, ACI online", "hit 2.4% vs base 2.2%, p=0.46 -> still not significant; diagnosis: 11-event low power + FR publication date lags market-moving announcement", "Closed-Negative"],
    ["G. Forward event eval (G3)", "Signals-approach EWS on 2024-07..2026-07 forward split, ACI flags, pre-forward-fixed shock threshold (rolling-2y sensitivity identical)", "1 shock onset (2025-02-26, Feb-2025 tariff reinstatement): flagged at onset (lead 0d), candidate proclamation surfaced (pub. 8d before); precision 3.6% honest false-alarm rate; n=1 -> case study, no rate claims", "Closed"],
    ["H. ACI + GARCH promoted to live pipeline", "src/pipeline.py production path: incremental band-history cache (parquet), ACI online calibration, GARCH(1,1) + changepoint cross-checks on alerts, attribution as candidate-surfacing", "Verified end-to-end via python -m src.pipeline; alerts carry garch_agrees / changepoint_agrees / ACI coverage", "Closed"],
    ["I. Announcement-dated retest (TPU)", "Daily TPU index (Caldara-Iacoviello JME 2020) spike onsets, validated vs known announcements; identical pre-registered design as step 4; GDELT implemented as fallback (API hard-throttles)", "PRIMARY p=0.85 (hit 6.8% vs base 9.0%); all sensitivity cells null. Causal attribution REJECTED under both event datings -> permanently ships as candidate-surfacing; no further sources tried (would be p-hacking)", "Closed-Negative"],
    ["J. Alert-precision filters", "Pre-specified F1 (changepoint agreement +/-7d) and F2 (breach run>=2), KLR persistence-filter practice; evaluated on the step-5 forward window", "F1: 55->6 flags, precision 3.6%->33.3%, NSR 0.84->0.08, recall preserved 1/1 -> live default ON. F2: dropped the only true positive (gap-style shock) -> OFF. n=1 caveat recorded", "Closed"],
]
cstat = {"Closed": MINT, "Tested": LIGHTER, "Closed-Negative": AMBER}
for i, row in enumerate(closed):
    er = r + 1 + i
    for j in range(3):
        c = ws.cell(row=er, column=1 + j, value=row[j]); c.font = F(9.5, bold=(j == 0))
        c.alignment = LEFT_T; c.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=4, end_row=er, end_column=5)
    c = ws.cell(row=er, column=4, value=row[3]); c.font = F(9.5, True); c.alignment = CENTER
    c.fill = fill(cstat.get(row[3], WHITE))
    for cc in range(1, 6):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    ws.row_dimensions[er].height = 26

r = r + 1 + len(closed) + 2
box(ws, r, 1, r, 5, "Open gaps carried forward (do not drift, close these)", AMBER_DK, WHITE, size=11)
gaps = [
    "Causal attribution: CLOSED as rejected (two pre-registered nulls: FR-dated p=0.46, TPU announcement-dated p=0.85). Attribution permanently ships as candidate-surfacing with a source link; no further event sources will be tried against this hypothesis (p-hacking).",
    "Alert precision: CLOSED. F1 changepoint-agreement filter live by default (NSR 0.84->0.08, recall preserved); F2 persistence filter evaluated harmful and disabled. Revisit both as more forward shocks accrue (current evidence is n=1).",
    "Remaining v2 items (not gaps): Moirai-2 covariate backbone, additional verticals via new YAML, streaming alert delivery.",
]
for i, t in enumerate(gaps):
    er = r + 1 + i
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=5)
    cell = ws.cell(row=er, column=1, value="•  " + t)
    cell.font = F(9.5); cell.alignment = LEFT_T; cell.border = BORDER_THIN
    ws.row_dimensions[er].height = max(24, 14 * (1 + len(t) // 115))
    if i % 2 == 1:
        for cc in range(1, 6):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

# ----------------------------------------------------------------------------
# SHEET 12 - V2 PLAN & PROGRESS
# ----------------------------------------------------------------------------
ws = wb.create_sheet("12. v2 Plan & Progress")
ws.sheet_view.showGridLines = False
set_widths(ws, {"A": 26, "B": 40, "C": 40, "D": 14})
title_block(ws, "v2 Plan & Progress", "Workstreams, decisions, and pre-registered acceptance criteria", "D")

r = 4
ws.cell(row=r, column=1, value="Workstreams").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Workstream", "Scope", "Acceptance criteria (pre-registered)", "Status"])
w2 = [
    ["W1. Covariate-aware forecasting",
     "Exogenous covariates via TimesFM 2.5 XReg (forecast_with_covariates, timesfm[xreg]). "
     "Covariate set fixed upfront: UUP (USD), CL=F (oil), HG=F (copper), all yfinance daily, "
     "no key. Future covariate values = carry-forward persistence (no lookahead).",
     "PRIMARY: mean ACI-calibrated interval width at matched 90% coverage, XReg arm vs v1 "
     "cached arm on identical HRC rolling origins. SECONDARY: MASE. XReg ships only if it "
     "improves the primary without degrading coverage; else v1 stands and Moirai-2 is evaluated.",
     "Done (negative)"],
    ["W2. Aluminum vertical",
     "config/aluminum.yaml only (Section 232 also covers aluminum); zero code changes.",
     "python -m src.pipeline --config config/aluminum.yaml runs end-to-end and produces a "
     "band-history cache + alert scan. Proves the extensibility NFR.",
     "Done"],
    ["W3. Alert delivery",
     "Pipeline writes machine-readable alert artifacts to outputs/alerts/<vertical>_<date>.json; "
     "optional webhook URL in config (stub, off by default).",
     "Artifact produced on a real run; schema documented; webhook config present but inert "
     "unless set.",
     "Done"],
]
sfill = {"In progress": AMBER, "Planned": LIGHTER, "Done": MINT,
         "Done (negative)": AMBER}
for i, row in enumerate(w2):
    er = r + 1 + i
    for j, val in enumerate(row):
        c = ws.cell(row=er, column=1 + j, value=val)
        c.font = F(9.5, bold=(j == 0)); c.alignment = LEFT_T; c.border = BORDER_THIN
    ws.cell(row=er, column=4).fill = fill(sfill.get(row[3], WHITE))
    ws.cell(row=er, column=4).alignment = CENTER
    ws.row_dimensions[er].height = 66

r = r + 1 + len(w2) + 2
box(ws, r, 1, r, 4, "W1 backbone decision (supersedes the v1-era 'Moirai-2 for v2' note)", BLUE, WHITE, size=11)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=4)
dec = ("DECISION: attempt covariates with TimesFM 2.5 XReg FIRST, keeping the proven v1 backbone, "
       "ACI calibration, and pipeline unchanged (one optional dependency: timesfm[xreg]). "
       "Moirai-2 remains the fallback, evaluated only if XReg fails to run or does not improve the "
       "pre-registered primary metric. Rationale: incremental risk, directly comparable arms on the "
       "same cached rolling origins, and no re-validation of the calibration stack.")
c = ws.cell(row=r, column=1, value=dec)
c.font = F(9.5); c.alignment = LEFT_T; c.fill = fill(LIGHT); c.border = BORDER_THIN

r += 5
ws.cell(row=r, column=1, value="v2 measured results").font = F(11, True, NAVY)
r += 1
table_header(ws, r, ["Workstream", "Result", "", "Status"])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
v2res = [
    ["W1 covariates (XReg)",
     "Pre-registered eval on identical HRC origins: XReg widened ACI intervals +22.7% at matched "
     "coverage (117.3 -> 143.9 USD/ton; coverage 90.0% vs 90.6%) and worsened MASE (6.85 -> 8.10). "
     "XReg does NOT ship; v1 stands. Moirai-2 not pursued: the failure is the persistence-covariate "
     "information content, not the backbone; revisit only with genuinely forward-informative "
     "covariates (futures curve, inventories). step8_covariates_eval.json",
     "Done (negative)"],
    ["W2 aluminum vertical",
     "config/aluminum.yaml (ALI=F) ran end-to-end with ZERO code changes: own band-history cache, "
     "event table, ACI scan, alert artifact. Extensibility NFR proven.",
     "Done"],
    ["W3 alert delivery",
     "deliver() writes outputs/alerts/<vertical>_<date>.json on every run (zero-alert runs "
     "distinguishable from no-run); candidate-drivers schema embeds the 'not a causal claim' note; "
     "webhook config-driven, inert by default, failures cannot kill the batch. Verified on steel "
     "and aluminum runs.",
     "Done"],
]
for i, row in enumerate(v2res):
    er = r + 1 + i
    c1 = ws.cell(row=er, column=1, value=row[0]); c1.font = F(9.5, True, NAVY)
    c1.alignment = LEFT_T; c1.border = BORDER_THIN
    ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=3)
    c2 = ws.cell(row=er, column=2, value=row[1]); c2.font = F(9.5); c2.alignment = LEFT_T
    c4 = ws.cell(row=er, column=4, value=row[2]); c4.font = F(9.5, True); c4.alignment = CENTER
    c4.fill = fill(sfill.get(row[2], WHITE))
    for cc in range(1, 5):
        ws.cell(row=er, column=cc).border = BORDER_THIN
    ws.row_dimensions[er].height = 62

r = r + 1 + len(v2res) + 2
box(ws, r, 1, r, 4, "v2 evaluation principles (carried from v1)", GREY, WHITE, size=10)
principles = [
    "Every claim pre-registered before results are seen; nulls are published, not buried.",
    "Arms compared on identical rolling origins with the same ACI settings (gamma=0.02, warmup=50).",
    "No lookahead: future covariate values are carry-forward persistence, never realized futures.",
    "Results land in this sheet + README + a step JSON before any status flips to Done.",
]
for i, t in enumerate(principles):
    er = r + 1 + i
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=4)
    c = ws.cell(row=er, column=1, value="•  " + t)
    c.font = F(9.5); c.alignment = LEFT_T; c.border = BORDER_THIN
    if i % 2 == 1:
        for cc in range(1, 5):
            ws.cell(row=er, column=cc).fill = fill(LIGHTER)

r = r + len(principles) + 2
box(ws, r, 1, r, 4, "Post-v2 gap register (found in the v2 close-out audit; solve in order)", AMBER_DK, WHITE, size=11)
r += 1
table_header(ws, r, ["Gap", "Risk if left open", "Fix", "Status"])
pv2 = [
    ["G1. No automated tests for core quant logic",
     "Regressions in CQR/ACI math, severity rules, attribution ranking, or the alert schema ship silently; all verification so far was manual",
     "pytest suite: synthetic coverage-guarantee tests for cqr_offset/aci_cqr (incl. regime-shift adaptation + negative-offset tightening), rule tests for severity, stub-embedder tests for attribution (incl. the object-dtype regression), event_eval metrics, deliver() schema + webhook-failure isolation. 25 tests green, network-free.",
     "Closed"],
    ["G2. Aluminum calibration unvalidated",
     "W2 proved the pipeline RUNS on ALI=F, not that ACI holds ~90% coverage there; alerts on an uncalibrated vertical are untrustworthy",
     "Validated (step9): ACI realized coverage 89.6% vs 90% target over 550 post-warmup days "
     "(2024-03..2026-07), 57 breach-days, mean calibrated width 300.7 USD/ton. Small-sample "
     "caveat noted; coverage continues to be monitored on every alert.",
     "Closed"],
    ["G3. ACI gamma=0.02 is an untested default",
     "A lucky constant: calibration quality could hinge on one literature value never checked on our series",
     "Swept (step10): coverage within ~0.5pp of target for gamma 0.005-0.05 on BOTH series "
     "(HRC 89.6-90.2%, SLX 89.2-89.7%); only gamma=0.1 degrades (>1pp). No challenger dominates; "
     "0.02 KEPT per the pre-registered rule. The default is robust, not lucky.",
     "Closed"],
    ["G4. Band cache lacks a config guard",
     "Changing horizon_days or the target symbol silently feeds stale cached bands into ACI (wrong intervals, no error)",
     "Implemented: sidecar bands_<vertical>.meta.json (symbol, horizon, quantiles) written with the "
     "cache; mismatch or missing meta invalidates and recomputes with a printed notice. 4 unit tests "
     "(hit, horizon change, symbol change, pre-guard cache). One-time recompute of existing caches.",
     "Closed"],
    ["G5. No scheduled execution",
     "The 'daily batch' only runs when someone remembers; stale alerts read as 'all clear'",
     "scripts/run_daily.bat for both verticals + README ops note with the Windows Task Scheduler registration command",
     "Open"],
]
gfill = {"Open": AMBER, "Closed": MINT, "Closed-Negative": AMBER}
for i, row in enumerate(pv2):
    er = r + 1 + i
    for j, val in enumerate(row):
        c = ws.cell(row=er, column=1 + j, value=val)
        c.font = F(9.5, bold=(j == 0)); c.alignment = LEFT_T; c.border = BORDER_THIN
    ws.cell(row=er, column=4).fill = fill(gfill.get(row[3], WHITE))
    ws.cell(row=er, column=4).alignment = CENTER
    ws.row_dimensions[er].height = 52

# ----------------------------------------------------------------------------
# Global: freeze panes on table sheets + set print/page basics
# ----------------------------------------------------------------------------
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr.fitToPage = True
    sh.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

import os
from pathlib import Path as _P
# Default: repo docs/ when run from tools/ inside the repo; else next to this script.
_here = _P(__file__).resolve().parent
_cand = _here.parent / "docs" / "Tariff_Shock_EWS_Solution_Design.xlsx"
_default = _cand if _cand.parent.exists() else _here / "Tariff_Shock_EWS_Solution_Design.xlsx"
OUT = os.environ.get("DESIGN_DOC_OUT", str(_default))
wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", [s.title for s in wb.worksheets])
